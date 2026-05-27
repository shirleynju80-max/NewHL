#!/usr/bin/env python3
"""
将 iFinD 导出的富时中国 A 股自由现金流聚焦指数（FCFQCD.FS）历史价格 xlsx 写入 index_bars.csv。

列映射（sheet「历史价格」）：
  交易日期 → date
  收盘价   → price_close、tri_close（富时仅 iFinD 这一条官方可用序列，二者同源）

元数据对照 factsheet（如 FCFQCDCH_YYYYMMDD.pdf）更新 indices.csv。

用法：
  python3 scripts/index_data_sync/import_ftse_fcfqcd_ifind_xlsx.py \\
    "/Users/.../com.51ifind富时中国A股自由现金流聚焦(FCFQCD.FS)-历史价格.xlsx"
"""
from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
INDEX_BARS = ROOT / "public" / "data" / "index_bars.csv"
INDICES = ROOT / "public" / "data" / "indices.csv"
INDEX_CODE = "FCFQCD"

FACTSHEET_META = {
    "name": "富时中国A股自由现金流聚焦指数",
    "market": "A",
    "category": "现金流",
    "methodology_summary": (
        "富时中国 A 股自由现金流聚焦指数（FTSE China A Free Cash Flow Focus），"
        "选股范围为富时中国 A 股自由流通指数，经自由现金流收益率与质量因子筛选约 50 只成分，"
        "指数层面目标高于基准的现金流收益率；单只成分权重上限 10%，季度审核（3/6/9/12 月）。"
        "行情唯一来源：iFinD 导出价格指数收盘（FCFQCD.FS，2013-12-31 起）；"
        "index_bars 中 tri_close 与 price_close 均使用该序列（无单独总收益日频数据）。"
    ),
    "methodology_url": (
        "https://www.lseg.com.cn/content/dam/ftse-russell/en_us/documents/"
        "ground-rules/ftse-china-a-free-cash-flow-focus-index-ground-rules-chinese.pdf"
    ),
    "inception_date": "2024-07-29",
    "base_date": "2013-12-31",
    "base_value": "1000",
    "launch_date": "2024-07-29",
    "weighting_method": "自由现金流比例加权；单只证券权重上限 10%",
    "rebalancing_frequency": "每季度（3、6、9、12 月）",
}


def normalize_date(raw) -> str:
    if raw is None:
        return ""
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise ValueError(f"无法识别日期: {raw}")


def parse_close(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().replace(",", "")
    if not s or s in {"-", "--", "—"}:
        return ""
    return f"{float(s):.4f}"


def read_index_bars() -> tuple[list[str], list[dict[str, str]]]:
    with INDEX_BARS.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        return fields, list(reader)


def write_index_bars(fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with INDEX_BARS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_ifind_rows(xlsx_path: Path) -> list[dict[str, str]]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("需要 openpyxl: python3 -m pip install openpyxl") from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = "历史价格" if "历史价格" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_date = headers.index("交易日期") if "交易日期" in headers else 0
    col_close = headers.index("收盘价") if "收盘价" in headers else 4

    out: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(r, col_date + 1).value
        if raw_date is None:
            continue
        date_s = str(raw_date).strip()
        if not date_s or date_s.startswith("数据"):
            break
        date = normalize_date(raw_date)
        close = parse_close(ws.cell(r, col_close + 1).value)
        if not close:
            continue
        out.append(
            {
                "index_code": INDEX_CODE,
                "date": date,
                "price_close": close,
                "tri_close": close,
                "div_yield_nominal_pct": "",
            }
        )
    wb.close()
    out.sort(key=lambda x: x["date"])
    return out


def merge_bars(imported: list[dict[str, str]]) -> int:
    fields, existing = read_index_bars()
    required = ["index_code", "date", "tri_close", "price_close", "div_yield_nominal_pct"]
    fieldnames = required + [x for x in fields if x not in required]
    kept = [row for row in existing if row.get("index_code") != INDEX_CODE]
    merged = imported + kept
    merged.sort(key=lambda x: (x.get("index_code", ""), x.get("date", "")))
    write_index_bars(fieldnames, merged)
    return len(imported)


def update_indices_meta() -> bool:
    if not INDICES.exists():
        print(f"skip indices update: {INDICES} not found")
        return False
    with INDICES.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    updated = False
    for row in rows:
        if row.get("index_code") != INDEX_CODE:
            continue
        for key, val in FACTSHEET_META.items():
            if key in row and row.get(key) != val:
                row[key] = val
                updated = True
            elif key in fieldnames:
                row[key] = val
                updated = True
        break
    else:
        print(f"warning: {INDEX_CODE} not in indices.csv")
        return False

    with INDICES.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return updated


def main() -> None:
    parser = argparse.ArgumentParser(description="导入 FCFQCD iFinD xlsx → index_bars + 更新 indices 元数据")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--skip-indices", action="store_true", help="仅写 index_bars，不更新 indices.csv")
    args = parser.parse_args()
    if not args.xlsx_path.is_file():
        raise SystemExit(f"文件不存在: {args.xlsx_path}")

    imported = load_ifind_rows(args.xlsx_path)
    if not imported:
        raise SystemExit("未解析到任何行情行")

    n = merge_bars(imported)
    print(f"index_bars: imported {n} rows for {INDEX_CODE}")
    print(f"  range {imported[0]['date']} .. {imported[-1]['date']}")
    print(f"  close sample last={imported[-1]['price_close']}")

    if not args.skip_indices:
        if update_indices_meta():
            print("indices.csv: updated FCFQCD metadata from factsheet")
        else:
            print("indices.csv: no changes")


if __name__ == "__main__":
    main()
