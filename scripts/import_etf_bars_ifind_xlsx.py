#!/usr/bin/env python3
"""
从 iFinD/腾讯财经导出的「ETF前复权历史行情」xlsx 写入 public/data/barsmore.csv。

每个 sheet 名为 ETF 代码；表头在第 4 行（B 列起）：日期、开盘价、收盘价、最高价、最低价、成交量。

用法：
  python3 scripts/import_etf_bars_ifind_xlsx.py \\
    "/Users/.../ETF前复权历史行情数据.xlsx" \\
    --codes 562080,560120,563990 \\
    --exclude-dates 2026-05-25
"""
from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
BARS_MORE = ROOT / "public" / "data" / "barsmore.csv"
FIELDS = ["etf_code", "date", "open", "high", "low", "close"]


def normalize_date(raw) -> str:
    if raw is None:
        return ""
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()[:10]
    if len(s) == 10 and s[4] == "-":
        return s
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def parse_price(raw) -> str:
    if raw is None:
        return ""
    s = str(raw).strip().replace(",", "")
    if not s or s in {"-", "--", "—"}:
        return ""
    return f"{float(s):.4f}"


def read_bars_csv(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    out: dict[tuple[str, str], dict[str, str]] = {}
    if not path.exists():
        return out
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = (row.get("etf_code") or "").strip()
            date = (row.get("date") or "").strip()
            if not code or not date:
                continue
            out[(code, date)] = {k: row.get(k, "") for k in FIELDS}
    return out


def write_bars_csv(path: Path, rows: dict[tuple[str, str], dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    sorted_keys = sorted(rows.keys(), key=lambda k: (k[0], k[1]))
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        for key in sorted_keys:
            writer.writerow(rows[key])


def load_sheet(ws, code: str, exclude: set[str]) -> list[dict[str, str]]:
    imported: list[dict[str, str]] = []
    for r in range(5, ws.max_row + 1):
        date = normalize_date(ws.cell(r, 2).value)
        if not date or date in exclude:
            continue
        o = parse_price(ws.cell(r, 3).value)
        c = parse_price(ws.cell(r, 4).value)
        h = parse_price(ws.cell(r, 5).value)
        lo = parse_price(ws.cell(r, 6).value)
        if not c:
            continue
        imported.append(
            {
                "etf_code": code,
                "date": date,
                "open": o or c,
                "high": h or c,
                "low": lo or c,
                "close": c,
            }
        )
    return imported


def main() -> None:
    parser = argparse.ArgumentParser(description="导入 ETF 前复权 xlsx → barsmore.csv")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument(
        "--codes",
        default="562080,560120,563990",
        help="逗号分隔 sheet 名 / ETF 代码",
    )
    parser.add_argument(
        "--exclude-dates",
        default="2026-05-25",
        help="剔除未收盘或无效日期，逗号分隔",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=BARS_MORE,
    )
    args = parser.parse_args()
    if not args.xlsx_path.is_file():
        raise SystemExit(f"文件不存在: {args.xlsx_path}")

    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit("需要 openpyxl: python3 -m pip install openpyxl") from e

    codes = [c.strip() for c in args.codes.split(",") if c.strip()]
    exclude = {d.strip() for d in args.exclude_dates.split(",") if d.strip()}

    merged = read_bars_csv(args.output)
    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    summary: list[str] = []
    for code in codes:
        if code not in wb.sheetnames:
            summary.append(f"{code}: sheet 不存在，跳过")
            continue
        ws = wb[code]
        rows = load_sheet(ws, code, exclude)
        replaced = 0
        for row in rows:
            key = (code, row["date"])
            if key in merged:
                replaced += 1
            merged[key] = row
        summary.append(
            f"{code}: +{len(rows)} 行"
            + (f"（覆盖同码同日 {replaced} 行）" if replaced else "")
            + (f"，区间 {rows[0]['date']}..{rows[-1]['date']}" if rows else "，无数据")
        )
    wb.close()

    write_bars_csv(args.output, merged)
    print(f"wrote {args.output} ({len(merged)} 行合计)")
    for line in summary:
        print(f"  {line}")
    if exclude:
        print(f"  已剔除日期: {', '.join(sorted(exclude))}")


if __name__ == "__main__":
    main()
